#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Mapeo de cargoId
cargo_map = {
    'Director Nivel Inicial': 2,
    'Director Nivel Secundaria': 3,
    'Director Nivel Primaria': 4,
    'Director Nivel Otros': 5,
    'Especialista Inicial - I1': 1,
    'Especialista Inicial - I2': 6,
    'Especialista Inicial - I3': 7,
    'Jefe de A.G.P': 8,
    'Director de UGEL': 9,
    'Docente de Primaria': 10
}

# Mapeo de distritos
distrito_map = {
    'Grocio Prado': 1,
    'Pueblo Nuevo': 2,
    'Alto Laran': 3,
    'Sunampe': 4,
    'Chincha Alta': 5,
    'El Carmen': 6,
    'San Juan de Yanac': 7,
    'San Pedro de Huacarpana': 8,
    'Chavin': 9,
    'Chincha Baja': 10,
    'Tambo de Mora': 11
}

# Mapeo de nivelModalidad
nivel_map = {
    'Inicial - Jardín': 'INICIAL-JARDIN',
    'Inicial-Jardin': 'INICIAL-JARDIN',
    'Primaria': 'PRIMARIA',
    'Secundaria': 'SECUNDARIA',
    'Técnico Productiva': 'EBA-CEPTPRO',
    'Básica Alternativa': 'EBA-CEPTPRO',
    'Básica Alternativa - Avanzado': 'EBA-CEPTPRO',
    'Básica Especial': 'EBA-CEPTPRO',
}

# Leer archivo de directivos
print("📖 Leyendo archivo de directivos...")
wb_origen = openpyxl.load_workbook(r'c:\Users\Yrvin PS\Desktop\DEVQORA\7 PROYECTO SISTEMA DE ASISTENCIA UGEL\BE-AsistenciaQRUgel\src\main\docs\-Lista-Directivos-QR (4).xlsx')
ws_origen = wb_origen.active

# Extraer datos
datos = []
for row_idx in range(2, ws_origen.max_row + 1):
    dni = ws_origen.cell(row_idx, 1).value
    nombres = ws_origen.cell(row_idx, 2).value
    apellidos = ws_origen.cell(row_idx, 3).value
    cargo_nombre = ws_origen.cell(row_idx, 4).value
    nivel = ws_origen.cell(row_idx, 10).value
    codigo_modular = ws_origen.cell(row_idx, 14).value
    distrito_nombre = ws_origen.cell(row_idx, 15).value
    
    # Convertir que puede ser None
    if not dni or not nombres or not apellidos:
        continue
    
    dni = str(dni).strip()
    nombres = str(nombres).strip()
    apellidos = str(apellidos).strip()
    
    # Mapear cargoId
    cargo_id = cargo_map.get(str(cargo_nombre).strip(), 1)
    
    # Mapear distritoId
    distrito_id = distrito_map.get(str(distrito_nombre).strip(), 1)
    
    # Mapear nivelModalidad
    nivel_modalidad = nivel_map.get(str(nivel).strip() if nivel else 'Primaria', 'PRIMARIA')
    
    # Buscar institucionEducativaId basado en código modular
    # Por ahora usamos un ID genérico (después se debe consultar BD)
    institucion_id = 1
    
    datos.append({
        'dni': dni,
        'nombres': nombres,
        'apellidos': apellidos,
        'cargoId': cargo_id,
        'distritoId': distrito_id,
        'nivelModalidad': nivel_modalidad,
        'institucionEducativaId': institucion_id,
        'estado': 'true'
    })

print(f"✅ Se extrajeron {len(datos)} directivos")

# Crear nuevo Excel
print("📝 Creando archivo de importación...")
wb_nuevo = Workbook()
ws_nuevo = wb_nuevo.active
ws_nuevo.title = "Personal"

# Headers
headers = ['dni', 'nombres', 'apellidos', 'cargoId', 'distritoId', 'nivelModalidad', 'institucionEducativaId', 'estado']
ws_nuevo.append(headers)

# Formatear headers
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
for cell in ws_nuevo[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Agregar datos
for fila in datos:
    ws_nuevo.append([
        fila['dni'],
        fila['nombres'],
        fila['apellidos'],
        fila['cargoId'],
        fila['distritoId'],
        fila['nivelModalidad'],
        fila['institucionEducativaId'],
        fila['estado']
    ])

# Ajustar ancho de columnas
ws_nuevo.column_dimensions['A'].width = 12  # dni
ws_nuevo.column_dimensions['B'].width = 18  # nombres
ws_nuevo.column_dimensions['C'].width = 22  # apellidos
ws_nuevo.column_dimensions['D'].width = 10  # cargoId
ws_nuevo.column_dimensions['E'].width = 12  # distritoId
ws_nuevo.column_dimensions['F'].width = 18  # nivelModalidad
ws_nuevo.column_dimensions['G'].width = 20  # institucionEducativaId
ws_nuevo.column_dimensions['H'].width = 10  # estado

# Guardar archivo
file_path = r'c:\Users\Yrvin PS\Desktop\DEVQORA\7 PROYECTO SISTEMA DE ASISTENCIA UGEL\BE-AsistenciaQRUgel\src\main\docs\ejemplo_personal_directivos.xlsx'
wb_nuevo.save(file_path)
print(f"✅ Archivo creado: {file_path}")
print(f"\n📊 Estadísticas:")
print(f"   - Total de directivos: {len(datos)}")
print(f"   - Por cargoId:")
for cargo, nombre in cargo_map.items():
    count = sum(1 for d in datos if d['cargoId'] == nombre)
    if count > 0:
        print(f"      {nombre}: {count}")
print(f"   - Por distritoId:")
for dist, id_n in sorted(distrito_map.items(), key=lambda x: x[1]):
    count = sum(1 for d in datos if d['distritoId'] == id_n)
    if count > 0:
        print(f"      {id_n} ({dist}): {count}")
