import openpyxl

# Abrir Excel
wb = openpyxl.load_workbook('Lista-Directivos-QR (3).xlsx')
ws = wb.active

# Mapping de niveles
nivel_map = {
    'Inicial - Jardín': 'INICIAL-JARDIN',
    'Inicial - Cuna Jardín': 'INICIAL-JARDIN',
    'Primaria': 'PRIMARIA',
    'Secundaria': 'SECUNDARIA',
    'Técnico Productivo': 'EBA-CEPTPRO',
    'Básica Alternativa': 'EBA-CEPTPRO',
    'Básica Alternativa - Avanzado': 'EBA-CEPTPRO',
    'Básica Especial': 'EBA-CEPTPRO',
    'Básica Especial - Primaria': 'EBA-CEPTPRO'
}

# Distritos
distritos = {
    'Grocio Prado': 1,
    'Pueblo Nuevo': 2,
    'Alto Laran': 3,
    'Sunampe': 4,
    'Chincha Alta': 5,
    'El Carmen': 6,
    'San Juan de Yanac': 7,
    'San Pedro de Huacarpana': 8,
    'Chavín': 9,
    'Chincha Baja': 10,
    'Tambo de Mora': 11
}

# Extraer datos
records = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if not row[0]:
        continue
    
    codigo_modular = row[1]
    nombre_ie = row[2]
    nivel = row[3]
    distrito_name = row[4]
    
    if not codigo_modular or not nombre_ie:
        continue
    
    nivel_mapped = nivel_map.get(str(nivel).strip() if nivel else '', 'PRIMARIA')
    distrito_id = distritos.get(str(distrito_name).strip() if distrito_name else '', 1)
    
    codigo_modular = str(codigo_modular).strip()
    nombre_ie = str(nombre_ie).strip().replace("'", "''")
    
    records.append((codigo_modular, nombre_ie, nivel_mapped, distrito_id))

# Generar SQL
sql = """-- Script para LLENAR la tabla INSTITUCIONES EDUCATIVAS
-- Ejecutar en el SQL Editor de Supabase en ambos schemas (produccion y certificacion)
-- Mapeo de Nivel/Modalidad:
--   Inicial - Jardín, Inicial - Cuna Jardín → INICIAL-JARDIN
--   Primaria → PRIMARIA
--   Secundaria → SECUNDARIA
--   Técnico Productivo, Básica Alternativa, Básica Especial → EBA-CEPTPRO

INSERT INTO "certificacion"."institucioneseducativas" (
  "codigoModular", "nombreIE", "nivelModalidad", "distritoId", "estado"
) 
VALUES 
"""

for i, (cod, nom, niv, dist) in enumerate(records):
    comma = "," if i < len(records) - 1 else ";"
    sql += f"  ('{cod}', '{nom}', '{niv}', {dist}, true){comma}\n"

sql += """
-- ═══════════════════════════════════════════════════════════════════════════════
-- Verificación: contar registros insertados
-- ═══════════════════════════════════════════════════════════════════════════════
-- SELECT COUNT(*) as total_instituciones, "distritoId", "nivelModalidad" 
-- FROM "certificacion"."institucioneseducativas"
-- GROUP BY "distritoId", "nivelModalidad"
-- ORDER BY "distritoId", "nivelModalidad";
"""

# Guardar archivo
with open('insert_todas_instituciones_educativas.sql', 'w', encoding='utf-8') as f:
    f.write(sql)

print(f'✅ Script generado: insert_todas_instituciones_educativas.sql')
print(f'Total registros: {len(records)}')
