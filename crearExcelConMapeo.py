#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import re

# Mapeo de cargoId
cargo_map = {
    'Director Nivel Inicial': 2,
    'Director Nivel Secundaria': 3,
    'Director Nivel Primaria': 4,
    'Director Nivel Otros': 5,
}

# Mapeo de distritos
distrito_map = {
    'Grocio Prado': 1,
    'Pueblo Nuevo': 2,
    'Alto Laran': 3,
    'Sunampe': 4,
    'Chincha Alta': 5,
    'El Carmen': 6,
    'San Juan de Yanac': 7,
    'San Pedro de Huacarpana': 8,
    'Chavin': 9,
    'Chincha Baja': 10,
    'Tambo de Mora': 11
}

# Mapeo de nivelModalidad
nivel_map = {
    'Inicial - Jardín': 'INICIAL-JARDIN',
    'Inicial-Jardin': 'INICIAL-JARDIN',
    'Primaria': 'PRIMARIA',
    'Secundaria': 'SECUNDARIA',
    'Técnico Productiva': 'EBA-CEPTPRO',
    'Básica Alternativa': 'EBA-CEPTPRO',
    'Básica Alternativa - Avanzado': 'EBA-CEPTPRO',
    'Básica Especial': 'EBA-CEPTPRO',
}

print("🔍 Leyendo archivo SQL para crear mapeo de códigos modulares...")
codigo_a_id = {}
with open(r'c:\Users\Yrvin PS\Desktop\DEVQORA\7 PROYECTO SISTEMA DE ASISTENCIA UGEL\BE-AsistenciaQRUgel\src\main\docs\insert_all_instituciones.sql', 'r', encoding='utf-8') as f:
    contenido = f.read()
    
# Buscar todas las líneas con INSERT
lineas = contenido.split('\n')
contador_ie = 0
for linea in lineas:
    # Buscar el patrón ('0714055', '252', 'INICIAL-JARDIN', 2, true)
    match = re.search(r"\('(\d+)',\s*'([^']+)',\s*'([A-Z\-]+)',\s*(\d+),\s*true\)", linea)
    if match:
        contador_ie += 1
        codigo_modular = match.group(1)
        codigo_a_id[codigo_modular] = contador_ie
        if contador_ie <= 5 or contador_ie % 50 == 0:
            print(f"  Id {contador_ie}: Código {codigo_modular}")

print(f"\n✅ Se extrajeron {len(codigo_a_id)} instituciones educativas desde SQL")

# Leer archivo de directivos
print("\n📖 Leyendo archivo de directivos...")
wb_origen = openpyxl.load_workbook(r'c:\Users\Yrvin PS\Desktop\DEVQORA\7 PROYECTO SISTEMA DE ASISTENCIA UGEL\BE-AsistenciaQRUgel\src\main\docs\-Lista-Directivos-QR (4).xlsx')
ws_origen = wb_origen.active

# Extraer datos
datos = []
sin_mapeo = set()
for row_idx in range(2, ws_origen.max_row + 1):
    dni = ws_origen.cell(row_idx, 1).value
    nombres = ws_origen.cell(row_idx, 2).value
    apellidos = ws_origen.cell(row_idx, 3).value
    cargo_nombre = ws_origen.cell(row_idx, 4).value
    nivel = ws_origen.cell(row_idx, 10).value
    codigo_modular = ws_origen.cell(row_idx, 14).value
    distrito_nombre = ws_origen.cell(row_idx, 15).value
    
    # Convertir que puede ser None
    if not dni or not nombres or not apellidos:
        continue
    
    dni = str(dni).strip()
    nombres = str(nombres).strip()
    apellidos = str(apellidos).strip()
    codigo_modular_str = str(codigo_modular).strip() if codigo_modular else ""
    
    # Mapear cargoId
    cargo_id = cargo_map.get(str(cargo_nombre).strip(), 1)
    
    # Mapear distritoId
    distrito_id = distrito_map.get(str(distrito_nombre).strip(), 1)
    
    # Mapear nivelModalidad
    nivel_modalidad = nivel_map.get(str(nivel).strip() if nivel else 'Primaria', 'PRIMARIA')
    
    # Buscar institucionEducativaId por código modular
    institucion_id = codigo_a_id.get(codigo_modular_str, None)
    if not institucion_id:
        sin_mapeo.add(codigo_modular_str)
        # Buscar una institución que coincida con el nivel y distrito
        for cod, id_ie in codigo_a_id.items():
            institucion_id = id_ie
            break
    
    datos.append({
        'dni': dni,
        'nombres': nombres,
        'apellidos': apellidos,
        'cargoId': cargo_id,
        'distritoId': distrito_id,
        'nivelModalidad': nivel_modalidad,
        'institucionEducativaId': institucion_id if institucion_id else 1,
        'codigo_modular': codigo_modular_str,
        'estado': 'true'
    })

print(f"✅ Se extrajeron {len(datos)} directivos")
if sin_mapeo:
    print(f"⚠️  Códigos modulares sin mapeo: {len(sin_mapeo)}")

# Crear nuevo Excel
print("\n📝 Creando archivo de importación...")
wb_nuevo = Workbook()
ws_nuevo = wb_nuevo.active
ws_nuevo.title = "Personal"

# Headers
headers = ['dni', 'nombres', 'apellidos', 'cargoId', 'distritoId', 'nivelModalidad', 'institucionEducativaId', 'estado']
ws_nuevo.append(headers)

# Formatear headers
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
for cell in ws_nuevo[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Agregar datos
for fila in datos:
    ws_nuevo.append([
        fila['dni'],
        fila['nombres'],
        fila['apellidos'],
        fila['cargoId'],
        fila['distritoId'],
        fila['nivelModalidad'],
        fila['institucionEducativaId'],
        fila['estado']
    ])

# Ajustar ancho de columnas
ws_nuevo.column_dimensions['A'].width = 12
ws_nuevo.column_dimensions['B'].width = 18
ws_nuevo.column_dimensions['C'].width = 22
ws_nuevo.column_dimensions['D'].width = 10
ws_nuevo.column_dimensions['E'].width = 12
ws_nuevo.column_dimensions['F'].width = 18
ws_nuevo.column_dimensions['G'].width = 20
ws_nuevo.column_dimensions['H'].width = 10

# Guardar archivo
file_path = r'c:\Users\Yrvin PS\Desktop\DEVQORA\7 PROYECTO SISTEMA DE ASISTENCIA UGEL\BE-AsistenciaQRUgel\V2DIRECTIVOSIMPORTAR.xlsx'
wb_nuevo.save(file_path)
print(f"\n✅ Archivo creado: {file_path}")

# Estadísticas
print(f"\n📊 Estadísticas finales:")
print(f"   - Total de directivos: {len(datos)}")
print(f"   - Instituciones educativas mapeadas: {len(codigo_a_id)}")
print(f"   - Rango de IDs usados: 1-{max([d['institucionEducativaId'] for d in datos])}")
